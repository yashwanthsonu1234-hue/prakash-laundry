from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from django.db.models import Count, Sum

import json

from .models import Order, Review
from .serializers import (
    OrderSerializer,
    RegisterSerializer,
    ReviewSerializer,
)

from collections import defaultdict
from collections import defaultdict



# ==========================
# DOWNLOAD PDF INVOICE
# ==========================
def download_invoice(request, order_id):
    order = Order.objects.get(id=order_id)

    try:
        services = json.loads(order.services)
    except:
        services = []

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = f'attachment; filename="invoice_{order.id}.pdf"'

    p = canvas.Canvas(response)

    y = 800

    p.setFont("Helvetica-Bold", 18)
    p.drawString(
        180,
        y,
        "PRAKASH LAUNDRY"
    )

    y -= 40

    p.setFont("Helvetica", 12)

    p.drawString(
        50,
        y,
        f"Invoice No: {order.id}"
    )

    y -= 25
    p.drawString(
        50,
        y,
        f"Customer: {order.customer_name}"
    )

    y -= 20
    p.drawString(
        50,
        y,
        f"Phone: {order.phone}"
    )

    y -= 20
    p.drawString(
        50,
        y,
        f"Address: {order.address}"
    )

    y -= 20
    p.drawString(
        50,
        y,
        f"Pickup Date: {order.pickup_date}"
    )

    y -= 20
    p.drawString(
        50,
        y,
        f"Status: {order.status}"
    )

    y -= 35

    p.setFont(
        "Helvetica-Bold",
        13
    )
    p.drawString(
        50,
        y,
        "Services"
    )

    y -= 25

    p.setFont(
        "Helvetica",
        12
    )

    for service in services:
        p.drawString(
            50,
            y,
            f"{service['title']} x {service['quantity']}"
        )

        p.drawString(
            350,
            y,
            f"Rs.{service['subtotal']}"
        )

        y -= 20

    y -= 15

    p.line(
        50,
        y,
        500,
        y
    )

    y -= 25

    p.setFont(
        "Helvetica-Bold",
        13
    )

    p.drawString(
        50,
        y,
        "Total Amount:"
    )

    p.drawString(
        350,
        y,
        f"Rs.{order.total_amount}"
    )

    y -= 50

    p.setFont(
        "Helvetica",
        12
    )

    p.drawString(
        120,
        y,
        "Thank You For Choosing"
    )

    y -= 20

    p.drawString(
        140,
        y,
        "Prakash Laundry Services"
    )

    p.save()

    return response


# ==========================
# ORDERS
# ==========================
class OrderCreateView(
    generics.ListCreateAPIView
):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer


class OrderUpdateView(
    generics.RetrieveUpdateDestroyAPIView
):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer


# ==========================
# REGISTER
# ==========================
class RegisterView(
    generics.CreateAPIView
):
    serializer_class = RegisterSerializer


# ==========================
# USER ORDERS
# ==========================
class UserOrderView(
    generics.ListCreateAPIView
):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Order.objects.filter(
            user=self.request.user
        )

    def perform_create(
        self,
        serializer
    ):
        serializer.save(
            user=self.request.user
        )


# ==========================
# TRACK ORDER
# ==========================
class TrackOrderView(
    APIView
):
    def get(
        self,
        request
    ):
        phone = request.GET.get(
            "phone"
        )

        if not phone:
            return Response(
                {
                    "error":
                    "Phone number required"
                },
                status=400,
            )

        orders = Order.objects.filter(
            phone=phone
        ).order_by(
            "-created_at"
        )

        serializer = OrderSerializer(
            orders,
            many=True
        )

        return Response(
            serializer.data
        )


# ==========================
# REVIEWS
# ==========================
class ReviewView(
    generics.ListCreateAPIView
):
    queryset = Review.objects.all().order_by(
        "-created_at"
    )
    serializer_class = ReviewSerializer


class ReviewDeleteView(
    generics.DestroyAPIView
):
    queryset = Review.objects.all()
    serializer_class = ReviewSerializer


class ReviewListView(
    APIView
):
    def get(
        self,
        request
    ):
        reviews = Review.objects.all()

        serializer = ReviewSerializer(
            reviews,
            many=True
        )

        return Response(
            serializer.data
        )


# ==========================
# DASHBOARD
# ==========================
class DashboardStatsView(
    APIView
):
    def get(
        self,
        request
    ):
        total_orders = Order.objects.count()

        pending_orders = Order.objects.filter(
            status="Pickup Scheduled"
        ).count()

        washing_orders = Order.objects.filter(
            status="Washing"
        ).count()

        delivered_orders = Order.objects.filter(
            status="Delivered"
        ).count()

        revenue = sum(
            order.total_amount
            for order in Order.objects.all()
        )

        return Response({
            "total_orders":
            total_orders,
            "pending_orders":
            pending_orders,
            "washing_orders":
            washing_orders,
            "delivered_orders":
            delivered_orders,
            "revenue":
            revenue,
        })

class AnalyticsView(APIView):

    def get(self, request):

        orders = Order.objects.all()

        revenue_by_month = defaultdict(float)

        service_count = defaultdict(int)

        for order in orders:

            month = order.created_at.strftime("%b")

            revenue_by_month[month] += float(
                order.total_amount
            )

            try:
                services = json.loads(
                    order.services
                )

                for service in services:
                    service_count[
                        service["title"]
                    ] += service["quantity"]

            except:
                pass

        revenue_data = []

        for month, revenue in revenue_by_month.items():

            revenue_data.append({
                "month": month,
                "revenue": revenue,
            })

        service_data = []

        for name, value in service_count.items():

            service_data.append({
                "name": name,
                "value": value,
            })

        return Response({
            "revenue_per_month": revenue_data,
            "top_services": service_data,
        })


class CustomerAnalyticsView(APIView):

    def get(self, request):

        customers = defaultdict(
            lambda: {
                "orders": 0,
                "total_spent": 0
            }
        )

        orders = Order.objects.all()

        for order in orders:

            phone = order.phone

            customers[phone]["name"] = (
                order.customer_name
            )

            customers[phone]["phone"] = (
                order.phone
            )

            customers[phone]["orders"] += 1

            customers[phone][
                "total_spent"
            ] += float(
                order.total_amount
            )

        result = []

        for customer in customers.values():

            customer["repeat_customer"] = (
                customer["orders"] > 1
            )

            result.append(customer)

        result.sort(
            key=lambda x: x["total_spent"],
            reverse=True
        )

        return Response(result)

def export_orders_excel(request):

    wb = Workbook()
    ws = wb.active

    ws.title = "Orders"

    ws.append([
        "Order ID",
        "Customer",
        "Phone",
        "Amount",
        "Status",
        "Pickup Date",
    ])

    for order in Order.objects.all():

        ws.append([
            order.id,
            order.customer_name,
            order.phone,
            order.total_amount,
            order.status,
            str(order.pickup_date),
        ])

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="orders.xlsx"'

    wb.save(response)

    return response

def export_customers_excel(request):

    wb = Workbook()
    ws = wb.active

    ws.title = "Customers"

    ws.append([
        "Customer",
        "Phone",
        "Orders",
        "Total Spent",
    ])

    customers = (
        Order.objects
        .values(
            "customer_name",
            "phone"
        )
        .annotate(
            orders=Count("id"),
            spent=Sum("total_amount")
        )
    )

    for c in customers:

        ws.append([
            c["customer_name"],
            c["phone"],
            c["orders"],
            c["spent"],
        ])

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="customers.xlsx"'

    wb.save(response)

    return response

def export_revenue_report(request):

    wb = Workbook()
    ws = wb.active

    ws.title = "Revenue"

    ws.append([
        "Month",
        "Revenue"
    ])

    revenue = (
        Order.objects
        .values("pickup_date__month")
        .annotate(
            revenue=Sum("total_amount")
        )
    )

    months = {
        1:"Jan",
        2:"Feb",
        3:"Mar",
        4:"Apr",
        5:"May",
        6:"Jun",
        7:"Jul",
        8:"Aug",
        9:"Sep",
        10:"Oct",
        11:"Nov",
        12:"Dec",
    }

    for r in revenue:

        ws.append([
            months[
                r["pickup_date__month"]
            ],
            r["revenue"]
        ])

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="revenue_report.xlsx"'

    wb.save(response)

    return response